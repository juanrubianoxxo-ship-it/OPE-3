"""Carga y normalización de las bases de datos de la aplicación.

``Operaciones_ult_semana.xlsm`` puede contener enlaces de mapa como texto,
hipervínculos de Excel o fórmulas ``HYPERLINK``. La carga conserva el destino
real del enlace, obtiene coordenadas en el orden latitud/longitud y registra la
fuente usada para que el mapa sólo reciba ubicaciones verificables.
"""
from __future__ import annotations

import base64
import os
import re
import tempfile
import unicodedata
from datetime import datetime, timezone

import requests
from collections import Counter
from typing import Any

import pandas as pd
import streamlit as st

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BOOK_PATH = os.path.join(BASE_DIR, "data", "Book.xlsx")
VISITAS_PATH = os.path.join(BASE_DIR, "data", "Operaciones_ult_semana.xlsm")
VISITAS_SHEET = "Visitas_Operaciones"
ONEDRIVE_URL_ENV = "ONEDRIVE_OPERACIONES_URL"
ONEDRIVE_URL_ALTERNATIVE = "ONEDRIVE_URL"
CACHE_TTL_SECONDS = 15 * 60

ESTADOS_VIGENTES = ["ABIERTA", "OBRA", "FIRMADA"]
MAPS_COLUMN_STD = "Enlace de la ubicación en Google Maps"

TIENDAS_COLS = [
    "NAME", "ESTADO", "PLAZA 2026", "DEPARTAMENTO", "MUNICIPIO",
    "UPZ/COMUNA", "ESTRATO", "TIPO DE LOCAL", "AREA", "X", "Y",
    "FECHA APE", "ARRENDADOR",
]

VISITAS_RENAME = {
    "Nombre del Punto ": "Nombre del Punto",
    " TICKET U6M": "TICKET U6M",
    " VENTAS OUM": "VENTAS OUM",
    " CONTRIBUCION UM": "CONTRIBUCION UM",
    " CONTRIBUCION U6M": "CONTRIBUCION U6M",
    " RENTA UM": "RENTA UM",
}

DATE_COLUMN_STD = "Fecha"
DATE_COLUMN_HINTS = ["fecha de visita", "fecha visita", "fecha de la visita", "fecha"]


def _get_configured_onedrive_url() -> str:
    """Obtiene el vínculo privado desde Secrets o variables de entorno."""
    configured = (
        os.getenv(ONEDRIVE_URL_ENV, "").strip()
        or os.getenv(ONEDRIVE_URL_ALTERNATIVE, "").strip()
    )
    try:
        secret_value = (
            str(st.secrets.get(ONEDRIVE_URL_ENV, "")).strip()
            or str(st.secrets.get(ONEDRIVE_URL_ALTERNATIVE, "")).strip()
        )
    except Exception:
        secret_value = ""
    return secret_value or configured


def _onedrive_share_to_direct(url: str) -> str:
    """Convierte vínculos compartidos en vínculos que fuerzan la descarga."""
    if not url:
        return url
    lowered = url.casefold()
    if "1drv.ms" in lowered or "download=1" in lowered:
        return url
    if "onedrive.live.com" in lowered or "sharepoint.com" in lowered:
        separator = "&" if "?" in url else "?"
        return f"{url}{separator}download=1"
    return url


def _onedrive_api_url(shared_url: str) -> str:
    """Convierte un vínculo compartido de OneDrive en URL de descarga."""
    encoded = base64.urlsafe_b64encode(shared_url.encode("utf-8")).decode("ascii").rstrip("=")
    return f"https://api.onedrive.com/v1.0/shares/u!{encoded}/root/content"


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _fetch_onedrive_file(shared_url: str) -> tuple[bytes, str, str]:
    """Descarga el libro usando el vínculo directo y devuelve sus metadatos."""
    if not shared_url:
        raise ValueError("No se configuró un vínculo de OneDrive.")

    direct_url = _onedrive_share_to_direct(shared_url)
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        )
    }
    candidates = [direct_url]
    # Fallback para vínculos compartidos que no aceptan ?download=1.
    if direct_url != shared_url:
        candidates.append(shared_url)
    candidates.append(_onedrive_api_url(shared_url))
    errores = []

    for candidate in candidates:
        try:
            response = requests.get(
                candidate, headers=headers, timeout=60, allow_redirects=True
            )
            response.raise_for_status()
            content_type = response.headers.get("content-type", "").casefold()
            if "text/html" in content_type or len(response.content) < 1000:
                raise ValueError(
                    "El vínculo devolvió una página HTML o un archivo demasiado pequeño. "
                    "Puede requerir inicio de sesión o no tener permisos públicos."
                )
            remote_modified = response.headers.get("last-modified", "")
            return response.content, remote_modified, response.url
        except Exception as exc:
            errores.append(f"{candidate[:100]}: {exc}")

    raise RuntimeError(
        "No fue posible descargar el archivo de OneDrive/SharePoint. "
        "Verifica que el vínculo permita 'Cualquier persona con el vínculo puede ver' "
        "y que apunte directamente al archivo .xlsm. Detalle: " + " | ".join(errores)
    )


def sync_operaciones_from_onedrive(shared_url: str | None = None, force: bool = False) -> dict:
    """Actualiza el XLSM local desde OneDrive y devuelve el estado de sincronización."""
    url = (shared_url or _get_configured_onedrive_url()).strip()
    if not url:
        return {"enabled": False, "source": "Archivo local", "message": "Sin vínculo de OneDrive configurado."}
    if force:
        _fetch_onedrive_file.clear()
    content, remote_modified, final_url = _fetch_onedrive_file(url)
    os.makedirs(os.path.dirname(VISITAS_PATH), exist_ok=True)
    previous = None
    if os.path.exists(VISITAS_PATH):
        with open(VISITAS_PATH, "rb") as current_file:
            previous = current_file.read()
    changed = previous != content
    if changed:
        fd, temp_path = tempfile.mkstemp(prefix="operaciones_", suffix=".xlsm", dir=os.path.dirname(VISITAS_PATH))
        try:
            with os.fdopen(fd, "wb") as temp_file:
                temp_file.write(content)
            os.replace(temp_path, VISITAS_PATH)
        finally:
            if os.path.exists(temp_path):
                os.unlink(temp_path)
    fetched_at = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    return {
        "enabled": True,
        "source": "OneDrive",
        "changed": changed,
        "bytes": len(content),
        "fetched_at": fetched_at,
        "remote_modified": remote_modified or "No informado por OneDrive",
        "url": final_url,
        "message": "Archivo actualizado desde OneDrive." if changed else "El archivo local ya estaba al día.",
    }


def get_onedrive_default_url() -> str:
    return _get_configured_onedrive_url()


def _file_signature(path: str) -> tuple[str, float, int]:
    """Devuelve metadatos que cambian cuando se reemplaza un libro."""
    stat = os.stat(path)
    return path, stat.st_mtime, stat.st_size


def _normalizar_encabezado(value: object) -> str:
    """Normaliza variantes de encabezados con espacios, tildes y guiones."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(text.casefold().strip().replace("_", " ").replace("-", " ").split())


def _find_date_column(df: pd.DataFrame) -> str | None:
    """Encuentra una columna de fecha en la hoja de operaciones."""
    cols_lower = {_normalizar_encabezado(c): c for c in df.columns if isinstance(c, str)}
    for hint in DATE_COLUMN_HINTS:
        normalized_hint = _normalizar_encabezado(hint)
        if normalized_hint in cols_lower:
            return cols_lower[normalized_hint]
    for normalized, original in cols_lower.items():
        if "fecha" in normalized:
            return original
    return None


def _find_maps_column(columns: list[object]) -> str | None:
    """Localiza la columna de enlaces aun cuando cambie levemente el título."""
    expected = _normalizar_encabezado(MAPS_COLUMN_STD)
    for column in columns:
        if _normalizar_encabezado(column) == expected:
            return column if isinstance(column, str) else None
    for column in columns:
        normalized = _normalizar_encabezado(column)
        if "enlace" in normalized and ("googlemaps" in normalized or "maps" in normalized):
            return column if isinstance(column, str) else None
    for column in columns:
        normalized = _normalizar_encabezado(column)
        if "maps" in normalized or "ubicacion" in normalized:
            return column if isinstance(column, str) else None
    return None


def _find_address_column(columns: list[object]) -> str | None:
    for column in columns:
        normalized = _normalizar_encabezado(column)
        if normalized in {"direccion", "address", "direcciondelpunto"} or "direccion" in normalized:
            return column if isinstance(column, str) else None
    return None


def _cell_to_text(value: object) -> str:
    """Convierte una celda en texto sin transformar vacíos en la palabra nan."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _formula_hyperlink_target(value: object) -> str:
    """Extrae el primer argumento de una fórmula HYPERLINK/HIPERVINCULO."""
    text = _cell_to_text(value)
    if not text or not re.match(r"^=\s*(?:HYPERLINK|HIPERVINCULO)\s*\(", text, flags=re.IGNORECASE):
        return ""
    match = re.search(r"^=\s*(?:HYPERLINK|HIPERVINCULO)\s*\(\s*\"([^\"]+)\"", text, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _target_from_excel_cell(cell: Any) -> str:
    """Obtiene el destino real de una celda hipervinculada de Excel."""
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is not None:
        target = _cell_to_text(getattr(hyperlink, "target", None))
        if target:
            return target
        location = _cell_to_text(getattr(hyperlink, "location", None))
        if location:
            return location
    formula_target = _formula_hyperlink_target(getattr(cell, "value", None))
    if formula_target:
        return formula_target
    raw_value = _cell_to_text(getattr(cell, "value", None))
    if raw_value.casefold().startswith(("https://", "http://", "www.")):
        return raw_value
    return ""


def _read_excel_hyperlink_targets() -> tuple[dict[int, str], str | None]:
    """Lee destinos de hipervínculos sin ejecutar macros del libro XLSM.

    Las claves del resultado son números de fila de Excel (la primera fila de
    datos es la 2). Se usa esa relación posicional para que IDs duplicados o
    formateados de manera distinta no asocien una ubicación a otro registro.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(
        VISITAS_PATH,
        read_only=False,
        data_only=False,
        keep_vba=True,
    )
    try:
        if VISITAS_SHEET not in workbook.sheetnames:
            return {}, None
        worksheet = workbook[VISITAS_SHEET]
        headers = [cell.value for cell in worksheet[1]]
        maps_header = _find_maps_column(headers)
        if maps_header is None:
            return {}, None
        maps_index = next(
            index for index, header in enumerate(headers, start=1)
            if _normalizar_encabezado(header) == _normalizar_encabezado(maps_header)
        )
        targets: dict[int, str] = {}
        for excel_row in range(2, worksheet.max_row + 1):
            target = _target_from_excel_cell(worksheet.cell(row=excel_row, column=maps_index))
            if target:
                targets[excel_row] = target
        return targets, maps_header
    finally:
        workbook.close()


def _add_visit_coordinates(df: pd.DataFrame) -> tuple[str | None, str | None]:
    """Ubica Operaciones exclusivamente con el enlace de Google Maps.

    No se usan las columnas X/Y, la dirección ni coordenadas previamente
    almacenadas en caché. Esto evita que un punto termine en una ubicación
    distinta a la del enlace que viene en ``Operaciones_ult_semana``.
    """
    maps_col = _find_maps_column(list(df.columns))
    address_col = _find_address_column(list(df.columns))

    if maps_col and maps_col != MAPS_COLUMN_STD:
        if MAPS_COLUMN_STD not in df.columns:
            df.rename(columns={maps_col: MAPS_COLUMN_STD}, inplace=True)
        else:
            df[MAPS_COLUMN_STD] = df[MAPS_COLUMN_STD].where(
                df[MAPS_COLUMN_STD].notna(), df[maps_col]
            )
        maps_col = MAPS_COLUMN_STD

    df["lat"] = pd.Series(float("nan"), index=df.index, dtype="float64")
    df["lon"] = pd.Series(float("nan"), index=df.index, dtype="float64")
    df["fuente_coordenadas"] = pd.Series("Sin enlace de Maps", index=df.index, dtype="object")
    df["enlace_maps_leido"] = pd.Series("", index=df.index, dtype="object")

    # Lee el destino real de hipervínculos, fórmulas HYPERLINK y texto visible.
    excel_targets: dict[int, str] = {}
    try:
        excel_targets, _ = _read_excel_hyperlink_targets()
    except Exception:
        excel_targets = {}

    if maps_col:
        visible_links = df[maps_col].map(_cell_to_text)
        hyperlink_links = df["_fila_excel"].map(excel_targets).fillna("").map(_cell_to_text)
        effective_links = hyperlink_links.where(hyperlink_links.ne(""), visible_links)
        df["enlace_maps_leido"] = effective_links
        df[maps_col] = effective_links

    # Cada registro se resuelve exclusivamente desde su link. La dirección se
    # envía vacía intencionalmente para impedir una ubicación aproximada.
    from src.maps_utils import get_coordinates_batch
    records = [
        (index, _cell_to_text(row.get("enlace_maps_leido", "")), "")
        for index, row in df.iterrows()
    ]
    for index, (latitude, longitude, source) in get_coordinates_batch(records).items():
        if latitude is not None and longitude is not None:
            df.at[index, "lat"] = latitude
            df.at[index, "lon"] = longitude
        df.at[index, "fuente_coordenadas"] = source

    valid = df["lat"].between(-90, 90) & df["lon"].between(-180, 180)
    df.loc[~valid, ["lat", "lon"]] = float("nan")
    df["ubicacion_mapeable"] = valid
    df["diagnostico_ubicacion"] = df["fuente_coordenadas"].where(
        valid, df["fuente_coordenadas"].replace("Sin enlace de Maps", "Sin coordenadas válidas del enlace")
    )
    return maps_col, address_col


@st.cache_data(show_spinner="Cargando tiendas vigentes...")
def load_tiendas(_sig=None) -> pd.DataFrame:
    _file_signature(BOOK_PATH)
    df = pd.read_excel(BOOK_PATH, sheet_name="JUN")
    df.columns = [str(c).strip() for c in df.columns]
    keep = [column for column in TIENDAS_COLS if column in df.columns]
    df = df[keep].copy()

    df["ESTADO"] = df["ESTADO"].astype(str).str.strip().str.upper()
    df = df[df["ESTADO"].isin(ESTADOS_VIGENTES)].copy()
    df["NAME"] = df["NAME"].astype(str).str.strip()
    df = df[df["NAME"].ne("") & df["NAME"].ne("0")]

    df["lat"] = pd.to_numeric(df["Y"], errors="coerce")
    df["lon"] = pd.to_numeric(df["X"], errors="coerce")
    valid = df["lat"].between(-90, 90) & df["lon"].between(-180, 180)
    df.loc[~valid, ["lat", "lon"]] = float("nan")
    return df.reset_index(drop=True)


@st.cache_data(show_spinner="Cargando puntos evaluados (Operaciones)...")
def load_visitas(_sig=None, include_coordinates: bool = True) -> pd.DataFrame:
    """Carga Operaciones; la geolocalización se puede omitir en vistas estadísticas."""
    _file_signature(VISITAS_PATH)
    df = pd.read_excel(VISITAS_PATH, sheet_name=VISITAS_SHEET, engine="openpyxl")
    df.columns = [str(column) for column in df.columns]
    df = df.rename(columns=VISITAS_RENAME)
    df.columns = [column.strip() if isinstance(column, str) else column for column in df.columns]
    # Relación estable con la fila del archivo para leer el hiperlink correcto.
    df["_fila_excel"] = df.index + 2

    if "Nombre del Punto" in df.columns:
        df["Nombre del Punto"] = df["Nombre del Punto"].astype(str).str.strip()
        df = df[df["Nombre del Punto"].ne("") & df["Nombre del Punto"].ne("nan")].copy()

    fecha_col = _find_date_column(df)
    if fecha_col is not None:
        df[DATE_COLUMN_STD] = pd.to_datetime(df[fecha_col], errors="coerce", dayfirst=True)
    else:
        df[DATE_COLUMN_STD] = pd.NaT

    if "ID" in df.columns:
        df["ID"] = df["ID"].astype(str).str.strip()
    elif "Nombre del Punto" in df.columns:
        df["ID"] = df["Nombre del Punto"].astype(str)
    else:
        df["ID"] = df.index.astype(str)

    if include_coordinates:
        maps_col, address_col = _add_visit_coordinates(df)
    else:
        maps_col, address_col = None, None
        df["lat"] = float("nan")
        df["lon"] = float("nan")
        df["fuente_coordenadas"] = "No calculadas en vista estadística"
        df["ubicacion_mapeable"] = False
        df["diagnostico_ubicacion"] = "No calculadas en vista estadística"
    summary = Counter(df["fuente_coordenadas"].fillna("Sin coordenadas"))
    df.attrs["coordinate_sources"] = {
        "maps_column": maps_col,
        "address_column": address_col,
        "con_xy_y_x": "Y" in df.columns and "X" in df.columns,
        "total_registros": int(len(df)),
        "mapeables": int(df["ubicacion_mapeable"].sum()),
        "fuentes": dict(summary),
    }
    return df.drop(columns=["_fila_excel"], errors="ignore").reset_index(drop=True)


def reload_all() -> None:
    """Fuerza la recarga de los libros y de las redirecciones de mapas."""
    load_tiendas.clear()
    load_visitas.clear()
    try:
        from src.maps_utils import (
            _verified_coordinate_cache,
            geocode_address,
            google_place_coordinates,
            resolve_map_link,
        )
        resolve_map_link.cache_clear()
        google_place_coordinates.cache_clear()
        geocode_address.cache_clear()
        _verified_coordinate_cache.cache_clear()
    except Exception:
        pass
